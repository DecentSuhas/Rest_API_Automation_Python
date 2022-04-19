"""
 This file has all the functions related to excel operation
"""
import openpyxl

from Referral.data.text import localization_dict

locale_text_list = []


def get_ref_text_from_excel(locale):
    wb = openpyxl.load_workbook(r"C:\Users\320052425\Desktop\Rest_API_Automation_Python\Referral\localization.xlsx")

    counter = 0
    sheet = wb.active
    row = sheet.max_row
    col = sheet.max_column
    global locale_text_list
    for i in range(1, col+1):
        cell = sheet.cell(row=1, column=i)
        if cell.value == locale:
            for k in range(2, row - 1):
                cell = sheet.cell(row=k, column=i)
                locale_text_list.append(cell.value)
    for key in localization_dict.keys():
        localization_dict[key] = locale_text_list[counter]
        counter = counter + 1


get_ref_text_from_excel("English")